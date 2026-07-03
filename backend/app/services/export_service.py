"""
Export Service — Generazione PDF/Excel per bilancio e libro giornale.

Riusa balance_sheet_service e journal_service per i dati; qui solo formattazione.
"""
import io
import uuid
from datetime import date
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models.accounting import Account
from app.models.journal import JournalEntry, JournalLine

STUDIO_UUID = uuid.UUID(settings.STUDIO_ID)


def _fmt_money(value: Decimal) -> str:
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


async def get_libro_giornale_dettaglio(
    db: AsyncSession,
    client_entity_id: uuid.UUID,
    fiscal_year_id: uuid.UUID,
) -> list[dict]:
    """Registrazioni posted con righe e conti, per l'export del libro giornale."""
    result = await db.execute(
        select(JournalEntry, JournalLine, Account.codice, Account.nome)
        .join(JournalLine, JournalLine.journal_entry_id == JournalEntry.id)
        .join(Account, JournalLine.account_id == Account.id)
        .where(
            JournalEntry.client_entity_id == client_entity_id,
            JournalEntry.fiscal_year_id == fiscal_year_id,
            JournalEntry.studio_id == STUDIO_UUID,
            JournalEntry.stato == "posted",
        )
        .order_by(JournalEntry.data_registrazione, JournalEntry.numero_registrazione)
    )
    rows = result.all()

    entries: dict[uuid.UUID, dict] = {}
    for entry, line, codice, nome in rows:
        if entry.id not in entries:
            entries[entry.id] = {
                "numero_registrazione": entry.numero_registrazione,
                "data_registrazione": entry.data_registrazione,
                "descrizione": entry.descrizione,
                "causale": entry.causale,
                "lines": [],
            }
        entries[entry.id]["lines"].append(
            {
                "codice": codice,
                "nome": nome,
                "dare": Decimal(str(line.dare)),
                "avere": Decimal(str(line.avere)),
                "descrizione": line.descrizione,
            }
        )
    return list(entries.values())


# ── PDF ──────────────────────────────────────────────────────────────────────

def _pdf_header(ragione_sociale: str, subtitle: str, anno: int) -> list:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleCustom", parent=styles["Title"], fontSize=16, spaceAfter=2
    )
    subtitle_style = ParagraphStyle(
        "SubtitleCustom", parent=styles["Normal"], fontSize=11, textColor=colors.grey
    )
    return [
        Paragraph(ragione_sociale, title_style),
        Paragraph(f"{subtitle} — Esercizio {anno}", subtitle_style),
        Spacer(1, 0.6 * cm),
    ]


def generate_bilancio_pdf(
    ragione_sociale: str,
    anno: int,
    stato_patrimoniale: dict,
    conto_economico: dict,
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    elements: list = _pdf_header(ragione_sociale, "Bilancio d'Esercizio", anno)

    def section_table(title: str, voci: list, totale: Decimal, negative: bool = False) -> None:
        elements.append(Paragraph(title, styles["Heading3"]))
        data = [["Codice", "Conto", "Importo"]]
        for v in voci:
            data.append([v["codice"], v["nome"], _fmt_money(v["saldo"])])
        data.append(["", "Totale", _fmt_money(totale)])
        table = Table(data, colWidths=[2.5 * cm, 9 * cm, 4 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f1f5f9")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 0.5 * cm))

    elements.append(Paragraph("Stato Patrimoniale", styles["Heading2"]))
    section_table("Attivo", stato_patrimoniale["attivo"]["voci"], stato_patrimoniale["attivo"]["totale"])
    section_table("Passivo", stato_patrimoniale["passivo"]["voci"], stato_patrimoniale["passivo"]["totale"])

    elements.append(Spacer(1, 0.4 * cm))
    elements.append(Paragraph("Conto Economico", styles["Heading2"]))
    section_table("Ricavi", conto_economico["ricavi"]["voci"], conto_economico["ricavi"]["totale"])
    section_table("Costi", conto_economico["costi"]["voci"], conto_economico["costi"]["totale"])

    risultato = conto_economico["utile_perdita"]
    label = "Utile d'Esercizio" if risultato >= 0 else "Perdita d'Esercizio"
    elements.append(Paragraph(f"<b>{label}: {_fmt_money(risultato)} €</b>", styles["Heading3"]))

    doc.build(elements)
    return buffer.getvalue()


def generate_giornale_pdf(
    ragione_sociale: str,
    anno: int,
    entries: list[dict],
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements: list = _pdf_header(ragione_sociale, "Libro Giornale", anno)

    data = [["N.", "Data", "Descrizione", "Conto", "Dare", "Avere"]]
    for e in entries:
        first = True
        for line in e["lines"]:
            data.append(
                [
                    str(e["numero_registrazione"]) if first else "",
                    e["data_registrazione"].strftime("%d/%m/%Y") if first else "",
                    e["descrizione"] if first else "",
                    f"{line['codice']} {line['nome']}",
                    _fmt_money(line["dare"]) if line["dare"] else "",
                    _fmt_money(line["avere"]) if line["avere"] else "",
                ]
            )
            first = False

    table = Table(data, colWidths=[1.2 * cm, 2.2 * cm, 5 * cm, 5.5 * cm, 2.3 * cm, 2.3 * cm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (4, 0), (5, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()


# ── Excel ────────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = "1E293B"


def _style_header_row(ws, row: int, ncols: int) -> None:
    from openpyxl.styles import PatternFill

    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = PatternFill(start_color=_HEADER_FILL, end_color=_HEADER_FILL, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 50)


def generate_bilancio_excel(
    ragione_sociale: str,
    anno: int,
    stato_patrimoniale: dict,
    conto_economico: dict,
) -> bytes:
    wb = Workbook()

    def write_section(ws, start_row: int, title: str, voci: list, totale: Decimal) -> int:
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
        row = start_row + 1
        ws.cell(row=row, column=1, value="Codice")
        ws.cell(row=row, column=2, value="Conto")
        ws.cell(row=row, column=3, value="Importo")
        _style_header_row(ws, row, 3)
        row += 1
        for v in voci:
            ws.cell(row=row, column=1, value=v["codice"])
            ws.cell(row=row, column=2, value=v["nome"])
            ws.cell(row=row, column=3, value=float(v["saldo"])).number_format = "#,##0.00"
            row += 1
        ws.cell(row=row, column=2, value="Totale").font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(totale)).font = Font(bold=True)
        ws.cell(row=row, column=3).number_format = "#,##0.00"
        return row + 2

    ws_sp = wb.active
    ws_sp.title = "Stato Patrimoniale"
    ws_sp.cell(row=1, column=1, value=f"{ragione_sociale} — Esercizio {anno}").font = Font(bold=True, size=14)
    row = write_section(ws_sp, 3, "Attivo", stato_patrimoniale["attivo"]["voci"], stato_patrimoniale["attivo"]["totale"])
    write_section(ws_sp, row, "Passivo", stato_patrimoniale["passivo"]["voci"], stato_patrimoniale["passivo"]["totale"])
    _autosize(ws_sp)

    ws_ce = wb.create_sheet("Conto Economico")
    ws_ce.cell(row=1, column=1, value=f"{ragione_sociale} — Esercizio {anno}").font = Font(bold=True, size=14)
    row = write_section(ws_ce, 3, "Ricavi", conto_economico["ricavi"]["voci"], conto_economico["ricavi"]["totale"])
    row = write_section(ws_ce, row, "Costi", conto_economico["costi"]["voci"], conto_economico["costi"]["totale"])
    label = "Utile d'Esercizio" if conto_economico["utile_perdita"] >= 0 else "Perdita d'Esercizio"
    ws_ce.cell(row=row, column=2, value=label).font = Font(bold=True)
    ws_ce.cell(row=row, column=3, value=float(conto_economico["utile_perdita"])).font = Font(bold=True)
    ws_ce.cell(row=row, column=3).number_format = "#,##0.00"
    _autosize(ws_ce)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_giornale_excel(
    ragione_sociale: str,
    anno: int,
    entries: list[dict],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Giornale"
    ws.cell(row=1, column=1, value=f"{ragione_sociale} — Libro Giornale {anno}").font = Font(bold=True, size=14)

    headers = ["N.", "Data", "Descrizione", "Causale", "Conto", "Nome Conto", "Dare", "Avere"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    row = 4
    for e in entries:
        for line in e["lines"]:
            ws.cell(row=row, column=1, value=e["numero_registrazione"])
            ws.cell(row=row, column=2, value=e["data_registrazione"].strftime("%d/%m/%Y"))
            ws.cell(row=row, column=3, value=e["descrizione"])
            ws.cell(row=row, column=4, value=e["causale"])
            ws.cell(row=row, column=5, value=line["codice"])
            ws.cell(row=row, column=6, value=line["nome"])
            ws.cell(row=row, column=7, value=float(line["dare"])).number_format = "#,##0.00"
            ws.cell(row=row, column=8, value=float(line["avere"])).number_format = "#,##0.00"
            row += 1

    _autosize(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
